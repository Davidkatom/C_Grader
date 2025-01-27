import os
import tkinter as tk
from tkinter import ttk

import openpyxl
from openpyxl.reader.excel import load_workbook
from tkinterdnd2 import DND_FILES, TkinterDnD

import UIGrader
from c_file_runner import run_file
from UIGrader import check_ass

root_dir = "C:/Grading/Now/ass5"
input = root_dir + "/Input.txt"
output_file = root_dir + "/Output.txt"
c_file = ""
checkbox_states = {}


# results = UIGrader.start_grading_process(root_dir)

def load_from_excel():
    results = []

    # Load the workbook and select the active sheet
    wb = load_workbook(root_dir + "/grading_results.xlsx")
    sheet = wb.active

    # Iterate over rows, skipping the header row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[4] == "100" or row[4] == 100:
            continue
        # Create a dictionary for each row and map column values
        result = {
            "c_file": row[0],  # Column A: File name
            "grade": row[4],  # Column E: Code grade
            "report": row[6],  # Column G: Report
            "c_code": row[8],  # Column I: C code
            "output": row[9]
        }
        results.append(result)

    return results


results = load_from_excel()
index = 0


def save():
    global root
    name = os.path.basename(c_file)

    file_path = root_dir + "/grading_results.xlsx"
    # Open the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Find the row with the given c_file
    row_to_update = None
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        # print(row[0])# Iterate over the first column (C File)
        if row[0] == name:
            row_to_update = row[0]
            break
    if row_to_update is None:
        print(name + " Not found")
        return
    # If the row is found, populate the cells
    if row_to_update is not None:
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == name:
                # Update the corresponding cells
                row[4].value = grade.get()
                row[6].value = bottom_text_area.get("1.0", "end")
                row[7].value = grade.get()
                break
    else:
        print(f"C File '{c_file}' not found in the Excel file.")

    # Save the workbook
    wb.save(file_path)
    # print(f"Excel file '{file_path}' updated successfully.")


def reset_fields():
    text_area.delete("1.0", "end")

    right_text_area.config(state='normal')
    bottom_text_area.config(state='normal')

    right_text_area.delete("1.0", "end")
    bottom_text_area.delete("1.0", "end")
    grade.delete(0, "end")  # Deletes all text from the `Entry`

    right_text_area.config(state='disabled')


def populate_result(result):
    global c_file
    print(result)

    c_file = root_dir + "/Processed/" + result.get("c_file")
    name = os.path.basename(result.get("c_file"))

    title_label.config(text=name)  # Update title with file name
    title_index.config(text=fr'{index}/{len(results)}')

    reset_fields()
    right_text_area.config(state='normal')

    text_area.insert("1.0", result.get("c_code"))
    right_text_area.insert("1.0", result.get("output") if result.get("output") is not None else "0")
    bottom_text_area.insert("1.0", result.get("report") if result.get("report") is not None else "0")
    grade.insert(0, result.get("grade") if result.get("grade") is not None else "0")

    right_text_area.config(state='disabled')

def forward():
    global results
    global index
    index += 1
    if index >= len(results):
        index = 0
    save()

    results = load_from_excel()
    populate_result(results[index])
    if results[index].get("grade") == 100:
        forward()


def backward():
    global index
    index -= 1
    if index < 0:
        index = len(results) - 1
    save()
    populate_result(results[index])


def load_text_from_file(filepath):
    global c_file
    c_file = filepath
    filepath = filepath.strip('{}')
    try:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
        text_area.edit_separator()  # Start undo block
        text_area.delete("1.0", tk.END)
        text_area.insert(tk.END, content)
        text_area.edit_separator()  # End undo block
        title_label.config(text=filepath.split('/')[-1])  # Update title with file name
    except Exception as e:
        print(f"Error loading file: {e}")


def drop_file(event):
    dropped_files = event.data
    if dropped_files:
        first_file = dropped_files.split(' ')[0]
        load_text_from_file(first_file)


def compile_and_run():
    global input, output_file
    input = input_file_entry.get()
    output_file = output_file_entry.get()
    print(f"Compiling and running with input: {input}, output: {output_file}")
    output = run_file(c_file, input)
    right_text_area.config(state='normal')
    right_text_area.delete("1.0", "end")
    right_text_area.insert("1.0", output)
    right_text_area.config(state='disabled')

    ### check
    grade, tests_failed = check_ass(output_file, output)

    bottom_text_area.delete("1.0", "end")
    bottom_text_area.insert("end", "Grade: " + str(grade) + "\n")
    bottom_text_area.insert("end", tests_failed)


def toggle_checkbox(option):
    checkbox_states[option] = not checkbox_states.get(option, False)
    print(f"{option} set to {checkbox_states[option]}")


# Create a DnD-enabled root window
root = TkinterDnD.Tk()
root.title("Ass checker")

# Configure main window grid
root.columnconfigure(0, weight=0)
root.columnconfigure(1, weight=1)
root.columnconfigure(2, weight=1)
root.rowconfigure(2, weight=1)

# Title Label for file name
title_label = tk.Label(root, text="", font=("Helvetica", 14), anchor="center")
title_label.grid(row=0, column=0, columnspan=3, sticky="ew", pady=5)


# Top Frame for buttons and input/output fields
top_frame = tk.Frame(root)
top_frame.grid(row=1, column=0, columnspan=3, sticky="ew")
top_frame.columnconfigure((0, 1, 2, 3), weight=1)

# Compile and Run Button
compile_button = tk.Button(top_frame, text="Compile and Run", command=compile_and_run)
compile_button.grid(row=0, column=0, padx=5, pady=5)

title_index = tk.Label(top_frame, text="", font=("Helvetica", 14), anchor="center")
title_index.grid(row=1, column=1, columnspan=3, sticky="ew", pady=5)

# Input and Output File Fields
input_file_label = tk.Label(top_frame, text="Input File:")
input_file_label.grid(row=1, column=0, sticky="e", padx=5)
input_file_entry = tk.Entry(top_frame, width=40)
input_file_entry.insert(0, input)
input_file_entry.grid(row=1, column=1, padx=5)

output_file_label = tk.Label(top_frame, text="Output File:")
output_file_label.grid(row=1, column=2, sticky="e", padx=5)
output_file_entry = tk.Entry(top_frame, width=40)
output_file_entry.insert(0, output_file)
output_file_entry.grid(row=1, column=3, padx=5)

# Left Frame with vertical checkboxes
topleft_frame = tk.Frame(root)
topleft_frame.grid(row=0, column=0, sticky="nsw")
backward_button = tk.Button(topleft_frame, text="Backward", command=backward)
backward_button.grid(row=0, column=0, padx=5, pady=5)

forward_button = tk.Button(topleft_frame, text="Forward", command=forward)
forward_button.grid(row=0, column=1, padx=5, pady=5)

save_button = tk.Button(topleft_frame, text="Save", command=save)
save_button.grid(row=1, column=0, padx=5, pady=5)

grade = tk.Entry(topleft_frame, width=5)
grade.grid(row=1, column=1, padx=5)

# Left Frame with vertical checkboxes
left_frame = tk.Frame(root)
left_frame.grid(row=2, column=0, sticky="nsw")

# Add checkboxes
checkbox_label = tk.Label(left_frame, text="Options:")
checkbox_label.grid(row=0, column=0, sticky="w")
checkbox_options = ["Line breakers", "Remove return 0", "No comments", "Not compiling"]
for i, option in enumerate(checkbox_options, start=1):
    state_var = tk.BooleanVar(value=False)
    checkbox_states[option] = state_var.get()
    checkbox = tk.Checkbutton(
        left_frame, text=option, variable=state_var, command=lambda opt=option: toggle_checkbox(opt)
    )
    checkbox.grid(row=i, column=0, sticky="w")

# Center Frame for scrollable text (editable)
center_frame = tk.Frame(root)
center_frame.grid(row=2, column=1, sticky="nsew")
center_frame.rowconfigure(0, weight=1)
center_frame.columnconfigure(0, weight=1)

text_area = tk.Text(center_frame, wrap="word", undo=True, autoseparator=False)
text_area.grid(row=0, column=0, sticky="nsew")

center_scrollbar = ttk.Scrollbar(center_frame, orient="vertical", command=text_area.yview)
center_scrollbar.grid(row=0, column=1, sticky="ns")
text_area.config(yscrollcommand=center_scrollbar.set)

# Right Frame for split scrollable text areas (non-editable)
right_frame = tk.Frame(root)
right_frame.grid(row=2, column=2, sticky="nsew")
right_frame.rowconfigure((0, 1), weight=1)
right_frame.columnconfigure(0, weight=1)

right_text_area = tk.Text(right_frame, wrap="word", state="normal", height=15)
right_text_area.grid(row=0, column=0, sticky="nsew")
right_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=right_text_area.yview)
right_scrollbar.grid(row=0, column=1, sticky="ns")
right_text_area.config(yscrollcommand=right_scrollbar.set)

bottom_text_area = tk.Text(right_frame, wrap="word", state="normal", height=10)
bottom_text_area.grid(row=1, column=0, sticky="nsew")
bottom_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=bottom_text_area.yview)
bottom_scrollbar.grid(row=1, column=1, sticky="ns")
bottom_text_area.config(yscrollcommand=bottom_scrollbar.set)

# Initialize right text areas
right_text_area.insert("1.0", "Right Text Area - Top")
right_text_area.config(state="disabled")
bottom_text_area.insert("1.0", "Right Text Area - Bottom")
bottom_text_area.config(state="normal")

# Enable Drag-and-Drop
text_area.drop_target_register(DND_FILES)
text_area.dnd_bind('<<Drop>>', drop_file)

# Add drag-and-drop to the input and output file entry fields
input_file_entry.drop_target_register(DND_FILES)
input_file_entry.dnd_bind('<<Drop>>', lambda e: input_file_entry.delete(0, tk.END) or input_file_entry.insert(0,
                                                                                                              e.data.strip(
                                                                                                                  '{}')))

output_file_entry.drop_target_register(DND_FILES)
output_file_entry.dnd_bind('<<Drop>>', lambda e: output_file_entry.delete(0, tk.END) or output_file_entry.insert(0,
                                                                                                                 e.data.strip(
                                                                                                                     '{}')))
populate_result(results[0])
root.mainloop()
